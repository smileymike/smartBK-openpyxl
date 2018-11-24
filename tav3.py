
import csv
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Side, Border


DEST_FILENAME = '/home/anonymous/Cashbooks/cashbookTaxYr2018-2019.xlsx'
LAST_TAX_YEAR_FILENAME = '/home/anonymous/Cashbooks/cashbookTaxYr2017-2018.xlsx'
SPACE_AND_CHECK_COL = 2
SPACE_AND_TOTAL_BOX = 2
MIN_TYPES_TRANSACTION = 11
CENTER = Alignment(horizontal='center')
ROW_SPACE_BEFORE_TOTAL = 2
FONT_BOLD = Font(name='Calibri', bold=True)
TOTAL_BORDER_LINES = Border(top=Side(style='medium'), bottom=Side(style='double'))

TURNOVER_AND_OTHER_INCOME = ['Sales', 'Reward Scheme']
OPERATIONAL_COST = ['Net']

# ABBV of "Starting Point For Catagories" of column
# starting column is column K
SPFC = 11

cashbook = load_workbook(DEST_FILENAME)
print("Cashbook opened")

receipts = cashbook['Cashbook Receipts']
payments = cashbook['Cashbook Payments']
dla = cashbook["Director's Loan Account"]
pla = cashbook["Profit & Loss Account"]

#print(get_column_letter(11) + " " + get_column_letter(receipts.max_column - SPACE_AND_CHECK_COL))
#print(get_column_letter(11) + " " + get_column_letter(payments.max_column - SPACE_AND_CHECK_COL))

transaction_type_dict = {}
check_desc = []

# add catagories from file
with open('/home/anonymous/Dropbox/repeated_transactions.csv', newline='') as csvfile:
	reader = csv.reader(csvfile)
	for row in reader:
		transaction_type_dict[row[0]] = row[1]
		check_desc.append(''.join(row[0]))

hold = ""

for i in check_desc:
	hold = hold + i + '|'
hold = hold[:-1]  # remove last '|'

first_two = re.compile(hold)

for row in range(6,receipts.max_row):		# Description Column
	if receipts.cell(column=3, row=row).value != None:
		mo = first_two.search(receipts.cell(column=3, row=row).value)
		if mo != None:
			receipts.cell(column=9, row=row, value=transaction_type_dict[mo.group()])


for row in range(6,payments.max_row):		# Description Column
	if payments.cell(column=3, row=row).value != None:
		mo = first_two.search(payments.cell(column=3, row=row).value)
		if mo != None:
			payments.cell(column=9, row=row, value=transaction_type_dict[mo.group()])

# find out last tax year DLA last balance entry
last_taxyear_cashbook = load_workbook(filename = LAST_TAX_YEAR_FILENAME, data_only=True)

active_dla_ws = last_taxyear_cashbook["Director's Loan Account"]

for row in active_dla_ws.iter_rows(min_row=6, min_col=6, max_col=6, max_row=active_dla_ws.max_row):
	for cell in row:
		if cell.value != None:
			last_balance = cell.value

# enter opening balance (unlinked to previous taxyear cashbook)
# should be like this: ='file:///home/anonymous/Cashbooks/cashbookTaxYr2017-2018.xlsx'#$'Director''s Loan Account'.F22
dla['F6'] = last_balance

# initalise counter
row_counter = 7

# Making current tax year entries from receipts rows
for row in receipts.iter_cols(min_row=6, min_col=9, max_col=9 , max_row=receipts.max_row-SPACE_AND_TOTAL_BOX):
	for cell in row:
		# initalise or reset a transaction data
		one_dla_transaction = []
		if cell.value == "Director’s Loan Account":
			one_dla_transaction.append(receipts['B' + str(cell.row)].value)
			one_dla_transaction.append(receipts['C' + str(cell.row)].value)
			one_dla_transaction.append(None)
			one_dla_transaction.append(receipts['D' + str(cell.row)].value)
			one_dla_transaction.append(None)
			balance = "=F{0}-C{1}+D{1}".format(row_counter-1,row_counter)
			row_counter = row_counter + 1
			one_dla_transaction.append(balance)

			dla.append(one_dla_transaction)

			# =F6-C7+D7 for balance calculation


for row in payments.iter_cols(min_row=6, min_col=9, max_col=9, max_row=payments.max_row-SPACE_AND_TOTAL_BOX):
	for cell in row:
		if cell.value == "Director’s Loan Account":
			print('B' + str(cell.row))
			print('C' + str(cell.row))
			print('D' + str(cell.row))

# Date Column
for row in dla.iter_rows(min_row=6, max_col=1, max_row=dla.max_row):
	for cell in row:
		cell.number_format = "DD/MM/YYYY"
		cell.alignment = CENTER

# Formating Paid, Due to Director and Balance
for col in range(3,1+dla.max_column):
	_= dla.column_dimensions[get_column_letter(col)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'

# Total calculation for "Paid" & "Due to Director"
max_transaction = dla.max_row

for col in range(3, 5):
	_ = dla.cell(column=col, row=max_transaction+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}7:{0}{1})".format(get_column_letter(col),max_transaction))


# Formating Totals on PAID, DUE TO DIRECTOR
paid_total = dla["C"+str(dla.max_row)]
due_to_director_total = dla["D"+str(dla.max_row)]

paid_total = dla["C"+str(dla.max_row)]
paid_total.number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
paid_total.font = FONT_BOLD
paid_total.border = TOTAL_BORDER_LINES

due_to_director_total = dla["D"+str(dla.max_row)]
due_to_director_total.number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
due_to_director_total.font = FONT_BOLD
due_to_director_total.border = TOTAL_BORDER_LINES

Turnover_And_Other_Income_Formula = '='
#print(Turnover_And_Other_Income_Formula)

for col in receipts.iter_cols(min_row=5, min_col=SPFC, max_col=receipts.max_column-SPACE_AND_CHECK_COL, max_row=5):
	for cell in col:
		for catagory in TURNOVER_AND_OTHER_INCOME:
			if cell.value == catagory:
				if Turnover_And_Other_Income_Formula != '=':
					Turnover_And_Other_Income_Formula = Turnover_And_Other_Income_Formula + '+'
#					print(Turnover_And_Other_Income_Formula)
				Turnover_And_Other_Income_Formula = Turnover_And_Other_Income_Formula + "$'Cashbook Receipts'.${0}${1}".format(cell.column,receipts.max_row)
#				print(Turnover_And_Other_Income_Formula)

Operational_Cost_Formula = '='
print(Operational_Cost_Formula)

for col in payments.iter_cols(min_row=5, min_col=4, max_col=payments.max_column-SPACE_AND_CHECK_COL, max_row=5):
	for cell in col:
		for catagory in OPERATIONAL_COST:
			if cell.value == catagory:
				if Operational_Cost_Formula != '=':
					Operational_Cost_Formula = Operational_Cost_Formula + '+'
					print(Operational_Cost_Formula)
				Operational_Cost_Formula = Operational_Cost_Formula + "$'Cashbook Payments'.${0}${1}".format(cell.column,payments.max_row)
				print(Operational_Cost_Formula)





pla['E5'] = Turnover_And_Other_Income_Formula

pla['E8'] = Operational_Cost_Formula


print("Profit & Loss Account Worksheet created - need to delete and re-enter '=' on linked formulaes to maket it work")


print("Director's Loan Account Worksheet created")

cashbook.save(DEST_FILENAME)
print("Cashbook closed")
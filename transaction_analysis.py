
# finding a set of unique transactions


import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEST_FILENAME = 'cashbookTaxYr2018-2019v1.xlsx'
SPACE_AND_CHECK_COL = 2
SPACE_AND_TOTAL_BOX = 2
MIN_TYPES_TRANSACTION = 10

cashbook = load_workbook(DEST_FILENAME)

print(cashbook.sheetnames)

receipts = cashbook['Cashbook Receipts']
payments = cashbook['Cashbook Payments']


print(get_column_letter(10) + " " + get_column_letter(receipts.max_column - SPACE_AND_CHECK_COL))
print(get_column_letter(10) + " " + get_column_letter(payments.max_column - SPACE_AND_CHECK_COL))

description_list_p = []
description_list_r = []

myset = set()

for row in range(6,receipts.max_row):		# Description Column
	if receipts.cell(column=2, row=row).value != None:
		description_list_r.append(receipts.cell(column=2, row=row).value)
		myset.add(receipts.cell(column=2, row=row).value)

for row in range(6,payments.max_row):		# Description Column
	if payments.cell(column=2, row=row).value != None:
		description_list_p.append(payments.cell(column=2, row=row).value)
		myset.add(payments.cell(column=2, row=row).value)

print(myset)

mylist = list(myset)
cleanlist = []

for each_description in mylist:
	splited_description = each_description.split()
	if splited_description[1] != 'ON': # remove "ON"
		first_two_words = splited_description[0] + " " + splited_description[1]
	else:
		first_two_words = splited_description[0]
	cleanlist.append(first_two_words)

mycleanset = set(cleanlist)

print(mycleanset)
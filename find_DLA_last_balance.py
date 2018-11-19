from openpyxl import load_workbook

LAST_TAX_YEAR_FILENAME = '/home/anonymous/Cashbooks/cashbookTaxYr2017-2018.xlsx'

wb = load_workbook(filename = LAST_TAX_YEAR_FILENAME, data_only=True)

print(wb.sheetnames)

active_dla_ws = wb["Director's Loan Account"]

for row in active_dla_ws.iter_rows(min_row=6, min_col=6, max_col=6, max_row=active_dla_ws.max_row):
	for cell in row:
		if cell.value != None:
			last_balance = cell.value

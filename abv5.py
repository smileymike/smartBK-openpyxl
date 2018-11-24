import csv
import re
from datetime import date

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter



# Build Cashbook and save it. 

DEST_FILENAME = '/home/anonymous/Cashbooks/cashbookTaxYr2018-2019.xlsx'
COMPANY_NAME = '<Company Name> Company'
TAX_YEAR_ENDED = 'Year Ended 31-03-2019 (Start 1st April 2018)'

# ABBV of "Starting Point For Catagories" of column
# starting column is column K
SPFC = 11

ROW_SPACE_BEFORE_TOTAL = 1
COLUMN_SPACE_BEFORE_TOTAL = 1
SPACE_AND_CHECK_COL = 2

FONT_BOLD = Font(name='Calibri', bold=True)
CENTER = Alignment(horizontal='center')
CENTER_AND_WRAP_TEXT = Alignment(horizontal="center", wrap_text=True)
TOTAL_BORDER_LINES = Border(top=Side(style='medium'), bottom=Side(style='double'))

FOUR_FIGURE_COLUMN = 9
FIVE_FIGURE_COLUMN = 10
SIX_FIGURE_COLUMN = 13
SEVEN_FIGURE_COLUMN = 15

TURNOVER_AND_OTHER_INCOME = ['Sales', 'Reward Scheme']

receipt_transaction_tally = 0
payment_transaction_tally = 0


try:
	cashbook = load_workbook(DEST_FILENAME)
	print("Cashbook opened")

except:
	print("no cashbook exists... creating new Cashbook")

	cashbook = Workbook()

	receipts = cashbook.active
	receipts.title = "Cashbook Receipts"
	receipts.sheet_view.zoomScale = 75


	# constructing the layout

	# headings
	receipts['A1'] = COMPANY_NAME
	receipts['A2'] = TAX_YEAR_ENDED
	receipts['A3'] = 'Cashbook Receipts'
	receipts['D4'] = 'Gross Amount'
	receipts['A5'] = 'Ref'		
	receipts['B5'] = 'Date'
	receipts['C5'] = 'Description'
	receipts['D5'] = 'Bank'
	receipts['E5'] = 'Cash'
	receipts['F5'] = 'Other'
	receipts['G5'] = 'VAT'
	receipts['H5'] = 'Net'
	receipts['I5'] = 'Analysis'

	# Bolds
	receipts['A1'].font = FONT_BOLD
	receipts['A2'].font = FONT_BOLD
	receipts['A3'].font = FONT_BOLD


	# Merge cells in Gross
	receipts.merge_cells('D4:G4')
	receipts['C4'].alignment = CENTER

	for col in receipts.iter_cols(min_row=4, max_col=9, max_row=5):
		for cell in col:
			cell.font = FONT_BOLD
			cell.alignment = CENTER

	receipts.column_dimensions['A'].width = 11
	receipts.column_dimensions['B'].width = 11
	receipts.column_dimensions['C'].width = 32
	receipts.column_dimensions['D'].width = FIVE_FIGURE_COLUMN
	receipts.column_dimensions['E'].width = FIVE_FIGURE_COLUMN
	receipts.column_dimensions['F'].width = FIVE_FIGURE_COLUMN
	receipts.column_dimensions['G'].width = FIVE_FIGURE_COLUMN
	receipts.column_dimensions['H'].width = SIX_FIGURE_COLUMN
	receipts.column_dimensions['I'].width = SIX_FIGURE_COLUMN
	receipts.column_dimensions['J'].width = SIX_FIGURE_COLUMN

	payments = cashbook.copy_worksheet(receipts)
	payments.title = 'Cashbook Payments'
	payments['A3'] = 'Cashbook Payments'
	payments['A5'] = 'No'
	payments.sheet_view.zoomScale = 75

	receipts_catagories = []

	# add catagories from file
	with open('receipts_catagories.csv', newline='') as csvfile:
		reader = csv.reader(csvfile)
		for row in reader:
			receipts_catagories.append(''.join(row))

	for col in range(SPFC,SPFC+len(receipts_catagories)):
		_ = receipts.cell(column=col, row=5, value="{0}".format(receipts_catagories[col-SPFC]))
		receipts.column_dimensions[get_column_letter(col)].width = FIVE_FIGURE_COLUMN

	receipts_check_col = SPFC+len(receipts_catagories)+COLUMN_SPACE_BEFORE_TOTAL
	receipts.cell(column=receipts_check_col, row=5, value="Check")
	receipts.column_dimensions[get_column_letter(receipts_check_col)].width = SIX_FIGURE_COLUMN

	for col in receipts.iter_cols(min_row=5, max_col=receipts.max_column, max_row=5):
		for cell in col:
			cell.font = FONT_BOLD
			cell.alignment = CENTER_AND_WRAP_TEXT

			# Add Director's Loan Account located column in Receipt for DLA Worksheet
			if cell.value == "Director’s Loan Account":
				DLA_COLUMN_IN_RECEIPTS = cell.column

	payments_catagories = []

	# Add catagories from file that orginates from FreeAgent website
	with open('payments_catagories.csv', newline='') as csvfile:
		reader = csv.reader(csvfile)
		for row in reader:
			payments_catagories.append(''.join(row))

	for col in range(SPFC,SPFC+len(payments_catagories)):
		_ = payments.cell(column=col, row=5, value="{0}".format(payments_catagories[col-SPFC]))

	payments_check_col = SPFC+len(payments_catagories)+COLUMN_SPACE_BEFORE_TOTAL
	payments.cell(column=payments_check_col, row=5, value="Check")
	payments.column_dimensions[get_column_letter(payments_check_col)].width = SIX_FIGURE_COLUMN

	for col in payments.iter_cols(min_row=5, min_col=SPFC, max_col=payments.max_column, max_row=5):
		for cell in col:
			cell.font = FONT_BOLD
			cell.alignment = CENTER_AND_WRAP_TEXT

			# Add Director's Loan Account located column in Payments for DLA Worksheet
			if cell.value == "Director’s Loan Account":
				DLA_COLUMN_IN_PAYMENTS = cell.column

	receipts.row_dimensions[5].height = 40

	payments.row_dimensions[5].height = 40

	transactions = []

	# Read data.csv from Barclay Bank Website (download first)
	with open('/home/anonymous/Dropbox/data.csv', newline='') as csvfile:
		readCSV = csv.reader(csvfile, delimiter=',')
		for row in readCSV:
			if row[3] != 'Amount':
				del row[0]
				del row[1]
				del row[2]

				row[1], row[2] = row[2], row[1]

				row[1] = re.sub(r'\d\d\d\d\d\d(\d\d)?','',row[1])
				row[1] = row[1].replace("********************** ", "")
				row[1] = re.sub(' +',' ',row[1])

				words = row[1].split()

				if words[len(words)-1] == 'STO':
					del words[-1]
				elif words[len(words)-1] == 'ASD':
					del words[-1]
				elif words[len(words)-1] == 'BBP':
					del words[-1]
				elif words[len(words)-1] == 'CLP':
					del words[-1]
				elif words[len(words)-1] == 'BDC':
					del words[-1]
				elif words[len(words)-1] == 'FT':
					del words[-1]

				row[1] = ' '.join(words)

				# converting str date to date object
				entry_date = row[0].split("/")
				row[0] = date(int(entry_date[2]),int(entry_date[1]),int(entry_date[0]))
				row[2] = float(row[2])

				prefix_row = [' '] + row

				transactions.append(prefix_row)	

	for each_transaction in reversed(transactions):
		if each_transaction[3] > 0.00:
			receipt_transaction_tally += 1
			receipts.append(each_transaction)
		elif each_transaction[3] <= 0.00:
			# remove negative symbol
			each_transaction[3] = abs(each_transaction[3])
			payment_transaction_tally += 1
			payments.append(each_transaction)

	# Date Column
	for row in receipts.iter_rows(min_row=6, max_col=2,max_row=receipts.max_row):
		for cell in row:
			cell.number_format = "DD/MM/YYYY"
			cell.alignment = CENTER

	for row in payments.iter_rows(min_row=6, max_col=2,max_row=payments.max_row):
		for cell in row:
			cell.number_format = "DD/MM/YYYY"
			cell.alignment = CENTER

	# "Net" Amount Column
	for row in range(6,receipt_transaction_tally+6):
		_ = receipts.cell(column=8, row=row, value="=SUM(C{0}:E{0})-F{0}".format(row))

	for row in range(6,payment_transaction_tally+6):
		_ = payments.cell(column=8, row=row, value="=SUM(C{0}:E{0})-F{0}".format(row))	

	# Gross Amount Columns - Bank, Cash, Other, VAT
	for col in range(4, 9):
		_ = receipts.cell(column=col, row=6+receipt_transaction_tally+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}6:{0}{1})".format(get_column_letter(col),5+receipt_transaction_tally))

	for col in range(4, 9):
		_ = payments.cell(column=col, row=6+payment_transaction_tally+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}6:{0}{1})".format(get_column_letter(col),5+payment_transaction_tally))

	# catagorisation formula for Gross Amount
	for row in range(6,6+receipt_transaction_tally):
		for col in range(SPFC,receipts.max_column-1):
			_ = receipts.cell(column=col, row=row, value='=IF({0}$5=${2}{1},${3}{1},"")'.format(get_column_letter(col),row,'I','H'))

	for row in range(6,6+payment_transaction_tally):
		for col in range(SPFC,payments.max_column-1):
			_ = payments.cell(column=col, row=row, value='=IF({0}$5=${2}{1},${3}{1},"")'.format(get_column_letter(col),row,'I','H'))

	# Total amounts under each catagorised items e.g. Sales, Drawings etc.
	for row in range(6,6+receipt_transaction_tally):
		for col in range(SPFC,receipts.max_column-1):
			_ = receipts.cell(column=col, row=6+receipt_transaction_tally+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}6:{0}{1})".format(get_column_letter(col),5+receipt_transaction_tally))

	for row in range(6,6+payment_transaction_tally):
		for col in range(SPFC,payments.max_column-1):
			_ = payments.cell(column=col, row=6+payment_transaction_tally+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}6:{0}{1})".format(get_column_letter(col),5+payment_transaction_tally))

	# Calculation for check or unchecked transactions
	for row in range(6,receipt_transaction_tally+6):
		_ = receipts.cell(column=receipts.max_column, row=row, value="=SUM({1}{0}:K{0})-H{0}".format(row,get_column_letter(receipts.max_column-2)))

	for row in range(6,payment_transaction_tally+6):
		_ = payments.cell(column=payments.max_column, row=row, value="=SUM({1}{0}:K{0})-H{0}".format(row,get_column_letter(payments.max_column-2)))

	# Total amount in check or unchecked transactions
	_ = receipts.cell(column=receipts.max_column, row=6+receipt_transaction_tally+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}6:{0}{1})".format(get_column_letter(receipts.max_column),5+receipt_transaction_tally,'H','G'))

	_ = payments.cell(column=payments.max_column, row=6+payment_transaction_tally+ROW_SPACE_BEFORE_TOTAL, value="=SUM({0}6:{0}{1})".format(get_column_letter(payments.max_column),5+payment_transaction_tally,'H','G'))

	# Formating Gross Amount & Catagories of Transactions
	for col in range(3,1+receipts.max_column):
		_= receipts.column_dimensions[get_column_letter(col)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'

	for col in range(3,1+payments.max_column):
		_= payments.column_dimensions[get_column_letter(col)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'

	# Formatting the total row
	for row in receipts.iter_cols(min_row=receipt_transaction_tally+6+ROW_SPACE_BEFORE_TOTAL, max_col=receipts.max_column, max_row=receipt_transaction_tally+6+ROW_SPACE_BEFORE_TOTAL):
		for cell in row:
			if cell.value != None:
				cell.number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
				cell.font = FONT_BOLD
				cell.border = TOTAL_BORDER_LINES

	# Formatting the total row
	for row in payments.iter_cols(min_row=payment_transaction_tally+6+ROW_SPACE_BEFORE_TOTAL, max_col=payments.max_column, max_row=payment_transaction_tally+6+ROW_SPACE_BEFORE_TOTAL):
		for cell in row:
			if cell.value != None:
				cell.number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
				cell.font = FONT_BOLD
				cell.border = TOTAL_BORDER_LINES

	for col in receipts.iter_cols(min_row=5, min_col=SPFC, max_col=receipts.max_column-SPACE_AND_CHECK_COL, max_row=5):
		for cell in col:
			heading = cell.value
			longest_width = FIVE_FIGURE_COLUMN
			for word in heading.split():
				if len(word) > longest_width:
					longest_width = len(word) + 3
				receipts.column_dimensions[cell.column].width = longest_width + 1

	for col in payments.iter_cols(min_row=5, min_col=SPFC, max_col=payments.max_column-SPACE_AND_CHECK_COL, max_row=5):
		for cell in col:
			heading = cell.value
			longest_width = FIVE_FIGURE_COLUMN
			for word in heading.split():
				if len(word) > longest_width:
					longest_width = len(word) + 3
				payments.column_dimensions[cell.column].width = longest_width + 1


# Build Director's Loan Account worksheet
dla = cashbook.create_sheet("Director's Loan Account")
dla.title = "Director's Loan Account"
dla['A1'] = COMPANY_NAME
dla['A1'].font = FONT_BOLD
dla['A2'] = TAX_YEAR_ENDED
dla['A2'].font = FONT_BOLD
dla['A3'] = "Director's Loan Account"
dla['A3'].font = FONT_BOLD
dla.sheet_view.zoomScale = 75

dla['A5'] = 'Date'
dla['B5'] = 'Description'
dla['C5'] = 'Paid'
dla['D5'] = 'Due to Director'
dla['F5'] = 'Balance'		
dla['A5'].font = FONT_BOLD
dla['B5'].font = FONT_BOLD
dla['C5'].font = FONT_BOLD
dla['D5'].font = FONT_BOLD
dla['F5'].font = FONT_BOLD	

dla.column_dimensions['A'].width = 11
dla.column_dimensions['B'].width = 32
dla.column_dimensions['C'].width = FIVE_FIGURE_COLUMN
dla.column_dimensions['D'].width = FIVE_FIGURE_COLUMN
dla.column_dimensions['E'].width = 3

dla['A5'].alignment = CENTER
dla['B5'].alignment = CENTER
dla['C5'].alignment = CENTER
dla['D5'].alignment = CENTER_AND_WRAP_TEXT
dla['F5'].alignment = CENTER

dla.row_dimensions[5].height = 30

dla['A6'] = date(int('2018'),int('4'),int('1'))
cell = dla['A6']
cell.number_format = 'DD/MM/YYYY'
dla['B6'] = 'Opening Balance'

# Build Profit & Loss Account worksheet
pla = cashbook.create_sheet("Profit & Loss Account")
pla.title = "Profit & Loss Account"
pla['A1'] = COMPANY_NAME
pla['A2'] = TAX_YEAR_ENDED
pla['A3'] = "Profit & Loss Account"
pla['A1'].font = FONT_BOLD
pla['A2'].font = FONT_BOLD
pla['A3'].font = FONT_BOLD
pla.sheet_view.zoomScale = 75

pla['B5'] = 'Turnover and other income'		# Sales + "RS" from Cashbook Receipts worksheet
pla['B6'] = 'Cost of Sales'					# Where from ? 
pla['B7'] = 'Gross Profit'					# Gross Profit = E5-E6
pla['B8'] = 'Operational Costs'				# From Cashbook Payments worksheet
pla['B9'] = 'Professional Costs'			# Why this is necessary?
pla['B10'] = 'Net Profit/(Loss)'			# = E7-E8-E9
pla['B11'] = 'Tax'
pla['B12'] = 'Profit/(Loss) after Tax'		# = E10-E11

pla['E6'] = 0			
pla['E7'] = '= E5-E6'
pla['E8'] = 0
pla['E9'] = 0
pla['E10'] = '= E7-E8-E9'
pla['E11'] = 0
pla['E12'] = '= E10-E11'


# Build Profit & Loss Account worksheet
bs = cashbook.create_sheet("Balance Sheet")
bs.title = "Balance Sheet"
bs['A1'] = COMPANY_NAME
bs['A2'] = TAX_YEAR_ENDED
bs['A3'] = "Balance Sheet"
bs['A1'].font = FONT_BOLD
bs['A2'].font = FONT_BOLD
bs['A3'].font = FONT_BOLD
bs.sheet_view.zoomScale = 75

bs['B5'] = 'Fixed Assets'
bs['C6'] = 'Computer Equipment'
bs['C7'] = 'Furniture'
bs['B8'] = 'Current Assets'
bs['C9'] = 'Stock'
bs['C10'] = 'Debtors'
bs['C11'] = 'Cash'
bs['B13'] = 'Total Assets'
bs['B15'] = 'Current Liabilities'
bs['C16'] = 'Suppliers'
bs['C17'] = 'PAYE/NI'
bs['C18'] = 'VAT'
bs['C19'] = "Director's Loan Account"
bs['C20'] = 'Corporation Tax'
bs['C21'] = 'Long Term Liabilities'
bs['C22'] = 'Bank Loan'
bs['C23'] = 'Deferred Tax'
bs['B25'] = 'Total Current Liabilities'
bs['B27'] = 'Net Assets'
bs['B29'] = "Shareholder's Funds"
bs['C30'] = 'Share Capital'
bs['C31'] = 'Profit & Loss Account'
bs['B33'] = "Total Shareholder's Fund"
bs['B35'] = "Liabilities & Shareholder's Funds"

bs['B5'].font = FONT_BOLD
bs['B8'].font = FONT_BOLD
bs['B13'].font = FONT_BOLD
bs['B15'].font = FONT_BOLD
bs['B25'].font = FONT_BOLD
bs['B27'].font = FONT_BOLD
bs['B29'].font = FONT_BOLD
bs['B33'].font = FONT_BOLD
bs['B35'].font = FONT_BOLD

cashbook.save(DEST_FILENAME)
print("Cashbook closed")

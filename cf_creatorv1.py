from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import date, time, datetime, timedelta
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill


LAST_CASHFLOW_FORECAST_FILENAME = '/home/anonymous/dev/Python/SmartBK-openpyxl/Cunnane 13 Week Cashflow forecast-09-01-2019v1.xlsx'

FONT_BOLD = Font(name='Calibri', bold=True)
CENTER = Alignment(horizontal='center')
YELLOW_BKGROUND = PatternFill(bgColor="FFF200", fill_type = "gray0625")
TOTAL_BORDER_LINES = Border(top=Side(style='medium'), bottom=Side(style='double'))

today = date.today()
day = today.strftime("%d")
month = today.strftime("%b")
month_no = today.strftime("%m")
year = today.strftime("%Y")

dest_filename = 'Cunnane 13wk Cashflow Forecast-{0}-{1}-{2}.xlsx'.format(day,month,year)

COMPANY_NAME = '<Company Name> Company'
STARTDATE_HEADING = '{0}-{1}-{2}'.format(day,month[:3],year)

cashbook = Workbook()

cf = cashbook.active
cf.title = "Cashflow Forecast"
cf.sheet_view.zoomScale = 75

# constructing the layout
cf.freeze_panes = cf['D4']

# headings
cf['A1'] = COMPANY_NAME
cf['A2'] = STARTDATE_HEADING
cf['A3'] = '13 Weeks Cashflow Forecast'

cf['A1'].font = FONT_BOLD
cf['A2'].font = FONT_BOLD
cf['A3'].font = FONT_BOLD

cf.column_dimensions['A'].width = 3
cf.column_dimensions['B'].width = 3
cf.column_dimensions['C'].width = 29

#print(today.weekday())

# if today is not Sunday
if today.weekday() != 6:
	# work out date of first day of the week (Sunday is first day of the week)
	day = str(today)
	dt = datetime.strptime(day, '%Y-%m-%d')
	start = dt - timedelta(days=dt.weekday()+1)   # +1 is Sunday, omit it is Monday

	st_day = start.strftime("%d")
	st_month_no = start.strftime("%m")
	st_year = start.strftime("%Y")

	today_date = date(int(st_year),int(st_month_no),int(st_day))
else:
	today_date = date(int(year),int(month_no),int(day))

# Date Row
#for col in cf.iter_cols(min_row=1, max_col=15, max_row=1):
#	for cell in col:
#		cell.number_format = "DD/MM/YYYY"
#		cell.alignment = CENTER


for x in range(1,14):
	cf.cell(row=2,column=3+x, value = x)

# enter a date with yellow background
cf['D3'] = today_date
cf['D3'].font = FONT_BOLD
cf['D3'].fill = YELLOW_BKGROUND
cf['D3'].alignment = CENTER


for col in range(4,17):

	cf.column_dimensions['{0}'.format(get_column_letter(col))].width = 13
	cf['{0}2'.format(get_column_letter(col))].alignment = CENTER
	cf['{0}2'.format(get_column_letter(col))].font = FONT_BOLD
	cf['{0}3'.format(get_column_letter(col))].number_format = "DD/MM/YYYY"
	if col > 4:

		# calculate weekly dates
		dt = today_date
		start = dt - timedelta(days=-7*(col-4))
		cf['{0}3'.format(get_column_letter(col))].value = start

		cf['{0}3'.format(get_column_letter(col))].alignment = CENTER
		cf['{0}3'.format(get_column_letter(col))].font = FONT_BOLD

prev_cb = load_workbook(LAST_CASHFLOW_FORECAST_FILENAME)

prev_cf = prev_cb['Cashflow Forecast']

for row in prev_cf.rows:
	for cell in row:
		if cell.value == "INCOME":
			income_loc = str(cell.column) + str(cell.row)
			cf[income_loc] = 'INCOME'
			cf[income_loc].font = FONT_BOLD

		if cell.value == "EXPENDITURE":
			expenditure_loc = str(cell.column) + str(cell.row)
			cf[expenditure_loc] = "EXPENDITURE"
			cf[expenditure_loc].font = FONT_BOLD
		if cell.value != None and cell.column == 'B':
#			print(str(cell.column) + str(cell.row) + " " + cell.value)
			cf[str(cell.column) + str(cell.row)] = cell.value
		if cell.value != None and cell.column == 'C':
#			print(str(cell.column) + str(cell.row) + " " + cell.value)
			cf[str(cell.column) + str(cell.row)] = cell.value
		if cell.value == "TOTAL INCOME":
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		elif cell.value == "Direct Costs":
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		elif cell.value == "TOTAL EXPENDITURE":
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.row == 3 and cell.value != None:
			str_date = str(cell.value)
			if str_date[:10] == str(today_date):
				# starting point for new cf
				sp_for_new_cf = cell.column

num_col_new_cf = column_index_from_string(sp_for_new_cf)

col_range = prev_cf['{0}:P'.format(sp_for_new_cf)]

difference = None

# "copy and paste" from old cashflow forecast to new one
for row in col_range:
	for cell in row:
		if cell.row > 5:
			if cell.value != None:
				string = str(cell.value)
				if string[:1] != "=":
					# calculate difference yet?
					if difference == None:
						# only one calculation is required
						difference = column_index_from_string(cell.column) - 4 # column 'D'
					new_col = column_index_from_string(cell.column) - difference
#					print(new_col)
					cf['{0}{1}'.format(get_column_letter(new_col), cell.row)] = cell.value
income_row = None
expenditure_row = None

for row in cf.rows:
	for cell in row:
		if cell.value == "INCOME":
			income_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.value == "TOTAL INCOME":
			income_ttl_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.value == "EXPENDITURE":
			expenditure_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.value == "TOTAL EXPENDITURE":
			expend_ttl_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.value == "Income less expenditure":
			i_l_exp_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.value == "Opening balance":
			open_b_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
		if cell.value == "Closing balance":
			close_b_row = cell.row
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD

for row in cf.rows:
	for cell in row:
		if cell.row == income_ttl_row - 1 and column_index_from_string(cell.column) > 3:
			pass		
		if cell.row == income_ttl_row and column_index_from_string(cell.column) > 3:
			cf[str(cell.column) + str(cell.row)].value = '=SUM({0}{1}:{0}{2})'.format(str(cell.column), str(cell.row-2),str(income_row+1))
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
			cf[str(cell.column) + str(cell.row)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
			cf[str(cell.column) + str(cell.row)].border = TOTAL_BORDER_LINES
		if cell.row == expend_ttl_row and column_index_from_string(cell.column) > 3:
			cf[str(cell.column) + str(cell.row)].value = '=SUM({0}{1}:{0}{2})'.format(str(cell.column), str(cell.row-2),str(expenditure_row+1))
			cf[str(cell.column) + str(cell.row)].font = FONT_BOLD
			cf[str(cell.column) + str(cell.row)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
			cf[str(cell.column) + str(cell.row)].border = TOTAL_BORDER_LINES
		if cell.row == i_l_exp_row  - 1 and column_index_from_string(cell.column) > 3:
			pass
		if cell.row == i_l_exp_row and column_index_from_string(cell.column) > 3:
			cf[str(cell.column) + str(cell.row)].value = '={0}{1}-{0}{2}'.format(str(cell.column), str(income_ttl_row),str(expend_ttl_row))
		if cell.row == open_b_row and column_index_from_string(cell.column) > 4:
			one_shift_left = column_index_from_string(cell.column) - 1
			cf[str(cell.column) + str(cell.row)].value = '={0}{1}'.format(get_column_letter(one_shift_left), str(close_b_row))
		if cell.row == close_b_row and column_index_from_string(cell.column) > 3:
			cf[str(cell.column) + str(cell.row)].value = '={0}{1}-{0}{2}'.format(str(cell.column), str(i_l_exp_row),str(open_b_row))

# Formating
for col in range(4,2+cf.max_column):
	_= cf.column_dimensions[get_column_letter(col)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'

cf['D' + str(open_b_row)].fill = YELLOW_BKGROUND
cf['D' + str(open_b_row)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'

check_col = cf.max_column+1
for row in range(income_row+1, cf.max_row):
	cf['{0}{1}'.format(get_column_letter(check_col),row)].value = "=SUM(D{0}:P{0})".format(row)
	if income_ttl_row == row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].border = TOTAL_BORDER_LINES
		cf['{0}{1}'.format(get_column_letter(check_col),row)].font = FONT_BOLD
	if income_ttl_row - 1 == row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].number_format = '* #,##0.00 ;-* #,##0.00 ;* -# ;@'
		cf['{0}{1}'.format(get_column_letter(check_col),row)].value = ""
	if income_ttl_row + 1 == row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].value = ""
	if income_ttl_row + 2 == row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].value = ""
	if income_ttl_row + 3 == row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].value = ""
	if income_ttl_row + 4 == row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].value = ""
	if row == expend_ttl_row-1:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].number_format = 'General'
	if row == expend_ttl_row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].border = TOTAL_BORDER_LINES
		cf['{0}{1}'.format(get_column_letter(check_col),row)].font = FONT_BOLD
	if row > expend_ttl_row:
		cf['{0}{1}'.format(get_column_letter(check_col),row)].value = ""	

cf['Q2'].value = "Total"
cf['Q2'].alignment = CENTER
cf['Q2'].font = FONT_BOLD

#print(income_row)
#print(income_ttl_row)
#print(expenditure_row)
#print(expend_ttl_row)
#print(cf.max_column)
#print(i_l_exp_row)
#print(open_b_row)
#print(close_b_row)

cashbook.save(dest_filename)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

CASHBOOKS_FOLDER_LOCATION = '/home/anonymous/Cashbooks/'
LAST_TAXYEAR_FILE = CASHBOOKS_FOLDER_LOCATION + '/cashbookTaxYr2017-2018.xlsx'
CURRENT_TAXYEAR_FILE = CASHBOOKS_FOLDER_LOCATION + '/cashbookTaxYr2018-2019.xlsx'
CURRENT_YEAR = 2019
FONT_BOLD = Font(name='Calibri', bold=True)

YEAR_STARTING_POINT = 5

# Open Current Taxyear
cashbook = load_workbook(CURRENT_TAXYEAR_FILE)
print('Cashbook opened again!')

receipts = cashbook['Cashbook Receipts']
payments = cashbook['Cashbook Payments']
dla = cashbook["Director's Loan Account"]

print(receipts['D26'])
print(receipts.max_row)
print(payments['D43'])
print(payments.max_row)

# current year balance sheet
bs = cashbook["Balance Sheet"]


bs['E4'] = 2019
bs['E4'].font = FONT_BOLD

bs['E6'] = 0
bs['E7'] = 0
bs['E9'] = 0
bs['E10'] = 0




bs['E11'] = "=$'Cashbook Receipts'.D{0}-$'Cashbook Payments'.D{1}+G11".format(receipts.max_row,payments.max_row) 
bs['E13'].value = '=SUM(E6:E11)'

bs['E16'] = 0
bs['E17'] = 0
bs['E18'] = 0

bs['E19'] = "=$'Director''s Loan Account'.D{0}-$'Director''s Loan Account'.C{0}+$'Balance Sheet'.G19".format(dla.max_row) 

bs['E20'] = 0
bs['E21'] = 0
bs['E22'] = 0
bs['E23'] = 0

bs['E25'] = '=SUM(E16:E23)'
bs['E27'] = '=E13-E25'

bs['E30'] = 1

bs['E31'] = "=$'Profit & Loss Account'.$E$12+G31"
bs['E33'] = '=SUM(E30:E31)'
bs['E35'] = '=E25+E33'


# Read from Last Tax Year
ly_cashbook = load_workbook(LAST_TAXYEAR_FILE, data_only=True)
print('Cashbook opened again!')

# last year balance sheet
ly_bs = ly_cashbook["Balance Sheet"]

#print(bs.max_column)
# copy previous bs columns to current tax year balance sheet
for col in range(YEAR_STARTING_POINT,ly_bs.max_column+1):
	if ly_bs['{0}4'.format(get_column_letter(col))].value != None:
#		print(ly_bs['{0}4'.format(get_column_letter(col))].value)
#		print('{0}4'.format(get_column_letter(col+2)))
		bs['{0}4'.format(get_column_letter(col+2))].value = ly_bs['{0}4'.format(get_column_letter(col))].value
		bs['{0}4'.format(get_column_letter(col+2))].font = FONT_BOLD
		for row in range(5,ly_bs.max_row+1):
			if ly_bs[get_column_letter(col)+str(row)].value != None:
#				print('{0} {1}'.format(row,ly_bs[get_column_letter(col)+str(row)].value))
#				print('{0}{1}'.format(get_column_letter(col+2),row))
				bs['{0}{1}'.format(get_column_letter(col+2),row)].value = ly_bs[get_column_letter(col)+str(row)].value



#print(bs.max_row)
cashbook.save(CURRENT_TAXYEAR_FILE)
print("Cashbook closed")

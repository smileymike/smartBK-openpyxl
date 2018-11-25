from openpyxl import Workbook

wb = Workbook()

cashbook = wb.active

Receipts = wb.create_sheet('Cashbook Receipts')

Receipts['B4'] = 100

Receipts['E4'] = 10

Balance_Sheet = wb.create_sheet('Balance Sheet')

Balance_Sheet['B5'].value = "=$'Cashbook Receipts'.$B$4+$'Cashbook Receipts'.$E$4"

wb.save('test.xlsx')

